"""Excel export service — generates .xlsx workbooks from simulation results."""

from __future__ import annotations

import io
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from api.models.scenario import Scenario
from api.models.simulation_result import SimulationResponse

# ── Column definitions ──────────────────────────────────────────────────────

_DATA_COLUMNS: list[tuple[str, str | None]] = [
    # (header label, number_format or None)
    ("Persona", None),
    ("Bal p10", '"$"#,##0'),
    ("Bal p25", '"$"#,##0'),
    ("Bal p50 (Median)", '"$"#,##0'),
    ("Bal p75", '"$"#,##0'),
    ("Bal p90", '"$"#,##0'),
    ("IRR p10", "0.0%"),
    ("IRR p25", "0.0%"),
    ("IRR p50 (Median)", "0.0%"),
    ("IRR p75", "0.0%"),
    ("IRR p90", "0.0%"),
    ("Prob. of Success", "0.0%"),
    ("Employee Contrib", '"$"#,##0'),
    ("Employer Contrib", '"$"#,##0'),
    ("Total Contrib", '"$"#,##0'),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
_LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")    # light blue
_SECTION_FILL = PatternFill("solid", fgColor="2E75B6")  # medium blue


def _pct(value: float) -> str:
    """Format a decimal fraction as a percentage string for display in the header."""
    return f"{value * 100:.1f}%"


def _vesting_label(vesting) -> str:  # noqa: ANN001
    t = vesting.type
    if t == "immediate":
        return "Immediate"
    if t == "cliff":
        return f"Cliff ({vesting.years} yr)"
    if t == "graded":
        grades = ", ".join(f"Yr {y}: {int(p*100)}%" for y, p in sorted(vesting.schedule.items()))
        return f"Graded ({grades})"
    return str(vesting)


def _eligibility_label(months: int) -> str:
    if months == 0:
        return "Immediate"
    return f"{months} month{'s' if months != 1 else ''}"


def _sanitize_filename(name: str) -> str:
    """Replace non-alphanumeric/space/dash characters then spaces for a safe filename."""
    safe = re.sub(r"[^\w\s-]", "_", name).strip().replace(" ", "_")
    return safe or "simulation"


# ── Main public function ─────────────────────────────────────────────────────


def generate_workbook(scenario: Scenario, simulation_result: SimulationResponse) -> bytes:
    """Build an .xlsx workbook from a scenario and simulation result.

    Layout (single sheet "Results"):
    1. Plan Design Summary header block
    2. Blank row
    3. Simulation Assumptions block
    4. Two blank rows
    5. Column header row (bold, dark fill)
    6. One data row per persona

    Returns raw bytes suitable for streaming as an HTTP response.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    row = 1  # 1-indexed current row tracker

    # ── 1. Plan Design Summary title ────────────────────────────────────────
    title_cell = ws.cell(row=row, column=1,
                         value=f"Plan Design Summary — {scenario.name}")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = _HEADER_FILL
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20
    # Merge across first two columns for visual breadth
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    # ── 2. Plan design fields ────────────────────────────────────────────────
    pd = scenario.plan_design
    plan_rows: list[tuple[str, str]] = [
        ("Plan Name", pd.name),
        ("Auto Enrollment", "Yes" if pd.auto_enroll_enabled else "No"),
        ("Default Deferral Rate", _pct(pd.auto_enroll_rate)),
        ("Auto Escalation", "Yes" if pd.auto_escalation_enabled else "No"),
        ("Annual Escalation Rate", _pct(pd.auto_escalation_rate)),
        ("Escalation Cap", _pct(pd.auto_escalation_cap)),
    ]
    # Match formula — one row per tier
    if pd.match_tiers:
        for i, tier in enumerate(pd.match_tiers, start=1):
            label = "Match Formula" if i == 1 else f"Match Formula (tier {i})"
            value = f"{tier.match_rate*100:.0f}% on first {tier.on_first_pct*100:.0f}%"
            plan_rows.append((label, value))
    else:
        plan_rows.append(("Match Formula", "None"))
    plan_rows += [
        ("Match Vesting", _vesting_label(pd.match_vesting)),
        ("Match Eligibility", _eligibility_label(pd.match_eligibility_months)),
        ("Core Contribution Rate", _pct(pd.core_contribution_pct)),
        ("Core Vesting", _vesting_label(pd.core_vesting)),
        ("Core Eligibility", _eligibility_label(pd.core_eligibility_months)),
    ]
    for label, value in plan_rows:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = _LABEL_FILL
        ws.cell(row=row, column=2, value=value)
        row += 1

    # ── 3. Blank row ─────────────────────────────────────────────────────────
    row += 1

    # ── 4. Simulation Assumptions section label ───────────────────────────────
    sec_cell = ws.cell(row=row, column=1, value="Simulation Assumptions")
    sec_cell.font = Font(bold=True, color="FFFFFF")
    sec_cell.fill = _SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    # Assumptions values — prefer scenario.overrides, fall back to defaults
    overrides = scenario.overrides
    inflation = overrides.inflation_rate if overrides and overrides.inflation_rate is not None else 0.025
    wage_growth = overrides.wage_growth_rate if overrides and overrides.wage_growth_rate is not None else 0.03

    assumption_rows: list[tuple[str, str]] = [
        ("Retirement Age", str(simulation_result.retirement_age)),
        ("Planning Age", str(simulation_result.planning_age)),
        ("Number of Simulations", str(simulation_result.num_simulations)),
        ("Inflation Rate", _pct(inflation)),
        ("Wage Growth Rate", _pct(wage_growth)),
    ]
    for label, value in assumption_rows:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = _LABEL_FILL
        ws.cell(row=row, column=2, value=value)
        row += 1

    # ── 5. Two blank rows before data table ───────────────────────────────────
    row += 2

    # ── 6. Column header row ──────────────────────────────────────────────────
    col_header_row = row
    for col_idx, (label, _fmt) in enumerate(_DATA_COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    # ── 7. Data rows — one per persona ────────────────────────────────────────
    if not simulation_result.personas:
        note_cell = ws.cell(row=row, column=1, value="No personas configured")
        note_cell.font = Font(italic=True, color="808080")
    else:
        for persona in simulation_result.personas:
            irr = persona.income_replacement_ratio
            rb = persona.retirement_balance
            total_contrib = (
                persona.total_employee_contributions + persona.total_employer_contributions
            )
            values: list = [
                persona.persona_name,
                rb.p10, rb.p25, rb.p50, rb.p75, rb.p90,
                irr.p10 if irr else "N/A",
                irr.p25 if irr else "N/A",
                irr.p50 if irr else "N/A",
                irr.p75 if irr else "N/A",
                irr.p90 if irr else "N/A",
                persona.probability_of_success,
                persona.total_employee_contributions,
                persona.total_employer_contributions,
                total_contrib,
            ]
            for col_idx, (value, (_, fmt)) in enumerate(
                zip(values, _DATA_COLUMNS), start=1
            ):
                cell = ws.cell(row=row, column=col_idx, value=value)
                if fmt and value != "N/A":
                    cell.number_format = fmt
            row += 1

    # ── 8. Auto column widths + freeze pane ──────────────────────────────────
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                try:
                    length = len(str(cell.value))
                except Exception:
                    length = 0
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze below the column header row so it stays visible when scrolling
    ws.freeze_panes = ws.cell(row=col_header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(scenario_name: str) -> str:
    """Return a safe .xlsx filename derived from the scenario name."""
    return f"{_sanitize_filename(scenario_name)}_results.xlsx"
