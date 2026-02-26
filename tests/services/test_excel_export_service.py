"""Unit tests for ExcelExportService."""

from io import BytesIO
from uuid import uuid4

import openpyxl
import pytest

from api.models.match_tier import MatchTier
from api.models.plan_design import PlanDesign
from api.models.scenario import Scenario
from api.models.simulation_result import PercentileValues, PersonaSimulationResult, SimulationResponse
from api.models.vesting import ImmediateVesting
from api.services.excel_export_service import generate_workbook


# ── Fixtures ─────────────────────────────────────────────────────────────────


def _make_plan_design() -> PlanDesign:
    return PlanDesign(
        name="Test 401(k)",
        match_tiers=[
            MatchTier(match_rate=1.0, on_first_pct=0.03),
            MatchTier(match_rate=0.5, on_first_pct=0.02),
        ],
        match_vesting=ImmediateVesting(),
        auto_enroll_enabled=True,
        auto_enroll_rate=0.06,
        auto_escalation_enabled=True,
        auto_escalation_rate=0.01,
        auto_escalation_cap=0.10,
        core_contribution_pct=0.02,
    )


def _make_scenario(plan_design: PlanDesign | None = None) -> Scenario:
    return Scenario(
        workspace_id=uuid4(),
        name="Test Scenario",
        plan_design=plan_design or _make_plan_design(),
    )


def _make_persona_with_irr(name: str = "Alice") -> PersonaSimulationResult:
    return PersonaSimulationResult(
        persona_id=uuid4(),
        persona_name=name,
        retirement_balance=PercentileValues(p10=400_000, p25=550_000, p50=700_000, p75=900_000, p90=1_100_000),
        income_replacement_ratio=PercentileValues(p10=0.60, p25=0.72, p50=0.85, p75=1.00, p90=1.20),
        probability_of_success=0.82,
        total_employee_contributions=120_000,
        total_employer_contributions=60_000,
        trajectory=[],
    )


def _make_persona_no_irr(name: str = "Bob") -> PersonaSimulationResult:
    return PersonaSimulationResult(
        persona_id=uuid4(),
        persona_name=name,
        retirement_balance=PercentileValues(p10=100_000, p25=150_000, p50=200_000, p75=250_000, p90=300_000),
        income_replacement_ratio=None,
        probability_of_success=0.30,
        total_employee_contributions=40_000,
        total_employer_contributions=20_000,
        trajectory=[],
    )


def _make_simulation_response(personas: list[PersonaSimulationResult]) -> SimulationResponse:
    return SimulationResponse(
        scenario_id=uuid4(),
        num_simulations=250,
        seed=42,
        retirement_age=67,
        planning_age=90,
        personas=personas,
    )


@pytest.fixture
def scenario() -> Scenario:
    return _make_scenario()


@pytest.fixture
def two_persona_result() -> SimulationResponse:
    return _make_simulation_response([
        _make_persona_with_irr("Alice"),
        _make_persona_no_irr("Bob"),
    ])


# ── Tests — core output validity ─────────────────────────────────────────────


class TestWorkbookValidity:
    def test_returns_bytes(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_parseable_as_xlsx(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        assert wb is not None

    def test_sheet_named_results(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        assert "Results" in wb.sheetnames


# ── Tests — data table row count ─────────────────────────────────────────────


class TestDataTableRows:
    def _get_all_values_in_col_a(self, wb) -> list:
        ws = wb.active
        return [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]

    def test_two_personas_produce_two_data_rows(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        col_a_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        # Count rows where column A matches a persona name
        persona_names = {p.persona_name for p in two_persona_result.personas}
        data_rows_found = sum(1 for v in col_a_values if v in persona_names)
        assert data_rows_found == 2

    def test_single_persona_produces_one_data_row(self, scenario):
        result_one = _make_simulation_response([_make_persona_with_irr("OnlyPerson")])
        result = generate_workbook(scenario, result_one)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        col_a_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert col_a_values.count("OnlyPerson") == 1


# ── Tests — column headers ───────────────────────────────────────────────────


class TestColumnHeaders:
    _EXPECTED_HEADERS = [
        "Persona", "Bal p10", "Bal p25", "Bal p50 (Median)", "Bal p75", "Bal p90",
        "IRR p10", "IRR p25", "IRR p50 (Median)", "IRR p75", "IRR p90",
        "Prob. of Success", "Employee Contrib", "Employer Contrib", "Total Contrib",
    ]

    def _find_col_header_row(self, ws) -> int | None:
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "Persona":
                return r
        return None

    def test_all_15_column_headers_present(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        header_row = self._find_col_header_row(ws)
        assert header_row is not None, "Column header row not found"
        headers = [ws.cell(row=header_row, column=c).value for c in range(1, 16)]
        assert headers == self._EXPECTED_HEADERS


# ── Tests — null IRR handling ────────────────────────────────────────────────


class TestNullIrrHandling:
    def _find_persona_row(self, ws, persona_name: str) -> int | None:
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == persona_name:
                return r
        return None

    def test_null_irr_renders_as_na(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        bob_row = self._find_persona_row(ws, "Bob")
        assert bob_row is not None, "Bob's row not found"
        # IRR columns are 7–11 (1-indexed)
        for col in range(7, 12):
            cell_val = ws.cell(row=bob_row, column=col).value
            assert cell_val == "N/A", f"Expected 'N/A' in IRR col {col}, got {cell_val!r}"

    def test_valid_irr_is_numeric(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        alice_row = self._find_persona_row(ws, "Alice")
        assert alice_row is not None
        for col in range(7, 12):
            cell_val = ws.cell(row=alice_row, column=col).value
            # openpyxl may read whole-number floats as int (e.g. 1.0 → 1)
            assert isinstance(cell_val, (int, float)), (
                f"Expected numeric in IRR col {col}, got {cell_val!r}"
            )


# ── Tests — total contributions calculation ──────────────────────────────────


class TestTotalContributions:
    def _find_persona_row(self, ws, persona_name: str) -> int | None:
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == persona_name:
                return r
        return None

    def test_total_contrib_equals_employee_plus_employer(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active

        alice_row = self._find_persona_row(ws, "Alice")
        assert alice_row is not None
        # Columns 13=Employee Contrib, 14=Employer Contrib, 15=Total Contrib
        emp = ws.cell(row=alice_row, column=13).value
        er = ws.cell(row=alice_row, column=14).value
        total = ws.cell(row=alice_row, column=15).value
        assert emp + er == pytest.approx(total)

        alice_persona = next(p for p in two_persona_result.personas if p.persona_name == "Alice")
        assert total == pytest.approx(
            alice_persona.total_employee_contributions + alice_persona.total_employer_contributions
        )


# ── Tests — plan design header block ────────────────────────────────────────


class TestPlanDesignHeader:
    def _col_a_values(self, ws) -> list:
        return [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]

    def test_title_row_contains_plan_design_summary(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        # Cell A1 should start with "Plan Design Summary"
        val = ws.cell(row=1, column=1).value
        assert val is not None and "Plan Design Summary" in str(val)

    def test_auto_enrollment_label_present_before_persona_header(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        col_a = self._col_a_values(ws)
        # Find position of "Persona" (column header row) and "Auto Enrollment"
        persona_idx = next((i for i, v in enumerate(col_a) if v == "Persona"), None)
        ae_idx = next((i for i, v in enumerate(col_a) if v == "Auto Enrollment"), None)
        assert persona_idx is not None and ae_idx is not None
        assert ae_idx < persona_idx, "Auto Enrollment label should appear before column header row"

    def test_simulation_assumptions_label_present(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        col_a = self._col_a_values(ws)
        assert "Simulation Assumptions" in col_a

    def test_match_formula_row_per_tier(self, scenario, two_persona_result):
        result = generate_workbook(scenario, two_persona_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        col_a = self._col_a_values(ws)
        match_rows = [v for v in col_a if v and "Match Formula" in str(v)]
        # Plan design has 2 match tiers → expect 2 Match Formula rows
        assert len(match_rows) == 2


# ── Tests — zero personas edge case ─────────────────────────────────────────


class TestZeroPersonas:
    def test_empty_personas_does_not_raise(self, scenario):
        empty_result = _make_simulation_response([])
        result = generate_workbook(scenario, empty_result)
        assert isinstance(result, bytes)
        wb = openpyxl.load_workbook(BytesIO(result))
        assert wb is not None

    def test_empty_personas_shows_note(self, scenario):
        empty_result = _make_simulation_response([])
        result = generate_workbook(scenario, empty_result)
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb.active
        all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "No personas configured" in all_values
