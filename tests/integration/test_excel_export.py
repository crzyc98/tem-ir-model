"""Integration tests for the Excel export endpoint."""

from io import BytesIO
from uuid import uuid4

import openpyxl
import pytest
from fastapi.testclient import TestClient

from api.main import create_app


# ── Shared helpers ────────────────────────────────────────────────────────────


def _create_workspace(client: TestClient) -> dict:
    resp = client.post("/api/v1/workspaces", json={"client_name": "ExportTestCo"})
    assert resp.status_code == 201
    return resp.json()


def _create_scenario(client: TestClient, workspace_id: str, name: str = "Export Test Plan") -> dict:
    resp = client.post(
        f"/api/v1/workspaces/{workspace_id}/scenarios",
        json={
            "name": name,
            "plan_design": {
                "name": "Standard 401(k)",
                "match_tiers": [
                    {"match_rate": 1.0, "on_first_pct": 0.03},
                    {"match_rate": 0.5, "on_first_pct": 0.02},
                ],
                "match_vesting": {"type": "immediate"},
                "auto_enroll_enabled": True,
                "auto_enroll_rate": 0.06,
                "auto_escalation_enabled": True,
                "auto_escalation_rate": 0.01,
                "auto_escalation_cap": 0.10,
                "core_contribution_pct": 0.0,
            },
        },
    )
    assert resp.status_code == 201
    return resp.json()


def _run_simulation(client: TestClient, workspace_id: str, scenario_id: str) -> dict:
    resp = client.post(
        f"/api/v1/workspaces/{workspace_id}/scenarios/{scenario_id}/simulate",
        json={},
    )
    assert resp.status_code == 200
    return resp.json()


# ── Test classes ─────────────────────────────────────────────────────────────


class TestExportEndpoint:
    """Core download — US1 acceptance criteria."""

    def test_export_returns_200_and_xlsx_content_type(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"])
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            assert resp.status_code == 200
            assert "spreadsheetml.sheet" in resp.headers["content-type"]

    def test_export_body_is_valid_xlsx(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"])
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            wb = openpyxl.load_workbook(BytesIO(resp.content))
            assert wb is not None
            assert "Results" in wb.sheetnames

    def test_export_has_content_disposition_attachment(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"])
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            cd = resp.headers.get("content-disposition", "")
            assert "attachment" in cd
            assert ".xlsx" in cd

    def test_export_404_for_nonexistent_scenario(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"])
            sim = _run_simulation(client, ws["id"], scenario["id"])

            fake_id = str(uuid4())
            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{fake_id}/export",
                json=sim,
            )
            assert resp.status_code == 404

    def test_export_404_for_nonexistent_workspace(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"])
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{uuid4()}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            assert resp.status_code == 404


class TestExportHeaderContent:
    """Plan design header — US2 acceptance criteria."""

    def test_cell_a1_contains_plan_design_summary(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"])
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            wb = openpyxl.load_workbook(BytesIO(resp.content))
            sheet = wb.active
            a1 = sheet.cell(row=1, column=1).value
            assert a1 is not None and "Plan Design Summary" in str(a1)

    def test_scenario_name_appears_in_header(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"], name="My Custom Plan")
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            wb = openpyxl.load_workbook(BytesIO(resp.content))
            sheet = wb.active
            first_few = " ".join(
                str(sheet.cell(row=r, column=1).value or "")
                for r in range(1, 5)
            )
            assert "My Custom Plan" in first_few

    def test_filename_reflects_scenario_name(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"], name="My Custom Plan")
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            cd = resp.headers.get("content-disposition", "")
            assert "My_Custom_Plan" in cd


class TestExportFilenameSpecialChars:
    """Filename sanitization — Polish T015."""

    def test_special_chars_in_scenario_name_produce_safe_filename(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"], name="Q4 Plan: 2026 Test!")
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            assert resp.status_code == 200
            cd = resp.headers.get("content-disposition", "")
            # Should not contain : or ! in the filename
            filename_part = cd.split("filename=")[-1].strip('"')
            assert ":" not in filename_part
            assert "!" not in filename_part
            assert filename_part.endswith(".xlsx")

    def test_special_chars_file_still_opens_successfully(self):
        app = create_app(base_path=None)
        with TestClient(app) as client:
            ws = _create_workspace(client)
            scenario = _create_scenario(client, ws["id"], name="Q4 Plan: 2026 Test!")
            sim = _run_simulation(client, ws["id"], scenario["id"])

            resp = client.post(
                f"/api/v1/workspaces/{ws['id']}/scenarios/{scenario['id']}/export",
                json=sim,
            )
            wb = openpyxl.load_workbook(BytesIO(resp.content))
            assert wb is not None
